import openpyxl
from openpyxl.styles import PatternFill

inp_file = openpyxl.open("Students.xlsx")
student_sheet = inp_file["Sheet1"]
student_per_school = {}
total_score_per_student = {}
print(student_sheet.max_row)
student_sheet.insert_cols(9)
student_sheet.cell(1, 9).value = "Total Score"
student_sheet.cell(1, 9).fill = PatternFill(fgColor="00FFFF00", patternType = "solid")
for row_count in range(2, student_sheet.max_row + 1):
    school_name = student_sheet.cell(row_count, 10).value
    #Find number of students in each school
    if school_name in student_per_school:
        student_count = student_per_school.get(school_name)
        student_per_school[school_name] = student_count + 1
    else:
        student_per_school[school_name] = 1
        print(f"First student of school {school_name}")
    score_total = 0

    #Find total mark for each student
    for score_per_sub in range(3, 9):
        score_total = score_total + student_sheet.cell(row_count, score_per_sub).value

    student_sheet.cell(row_count, 9).value = score_total
    student_name = student_sheet.cell(row_count, 2).value
    total_score_per_student[student_name] = score_total

print(total_score_per_student)
print(student_per_school)
inp_file.save("Students_marks.xlsx")
